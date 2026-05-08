"""File-format parsers — XLSX / CSV / JSON / YAML.

Returns a uniform `ParseResult` regardless of source format. All values
land as `str` (None → "") so downstream validators don't have to second-
guess type coercion. CSV delimiter (`;` vs `,`) is auto-detected from
the first non-empty line — Bitrix24/Excel exports default to `;` in
RU locales, native English exports default to `,`.

MAX_ROWS = 5000 is a hard ceiling. Anything above that returns a
ParseResult with `error` set; the caller (router) translates to HTTP
422 with a clear message. We don't truncate silently — the manager
needs to know they're missing data.
"""
from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from typing import Any

import yaml

from app.import_export.models import ImportJobFormat


MAX_ROWS = 5000


@dataclass
class ParseResult:
    headers: list[str] = field(default_factory=list)
    rows: list[dict[str, str]] = field(default_factory=list)
    row_count: int = 0
    error: str | None = None


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_file(
    content: bytes, filename: str, fmt: ImportJobFormat
) -> ParseResult:
    """Dispatch to the format-specific parser. Errors land in result.error
    (the router maps them to HTTP 422)."""
    try:
        if fmt == ImportJobFormat.xlsx:
            return _parse_xlsx(content)
        if fmt == ImportJobFormat.csv:
            return _parse_csv(content)
        if fmt == ImportJobFormat.json:
            return _parse_json(content)
        if fmt == ImportJobFormat.yaml:
            return _parse_yaml(content)
        return ParseResult(error=f"unsupported format: {fmt}")
    except Exception as exc:  # noqa: BLE001 — top-level guard, normalize to error
        return ParseResult(error=f"parse failed: {type(exc).__name__}: {exc}")


def detect_format(filename: str) -> ImportJobFormat | None:
    """Best-effort format detection from a filename extension. Used by
    the router when the client doesn't specify `?format=...`."""
    if "." not in filename:
        return None
    ext = filename.rsplit(".", 1)[-1].lower().strip()
    if ext == "xlsx":
        return ImportJobFormat.xlsx
    if ext == "csv":
        return ImportJobFormat.csv
    if ext == "json":
        return ImportJobFormat.json
    if ext in ("yaml", "yml"):
        return ImportJobFormat.yaml
    return None


# ---------------------------------------------------------------------------
# Format-specific helpers
# ---------------------------------------------------------------------------

def _coerce(value: Any) -> str:
    """Uniform string coercion. None → "", bools → 'true'/'false',
    everything else → str(...). Strips trailing whitespace."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _is_blank_row(row: dict[str, str]) -> bool:
    return not any(v for v in row.values())


def _from_dict_rows(headers_seed: list[str], items: list[Any]) -> ParseResult:
    """Shared helper for JSON / YAML formats: items must be list[dict]."""
    if not isinstance(items, list):
        return ParseResult(error="expected a list of objects")
    if len(items) > MAX_ROWS:
        return ParseResult(
            error=f"too many rows: {len(items)} > {MAX_ROWS}"
        )

    # Header set = union of keys across rows (preserving first-seen order)
    headers: list[str] = list(headers_seed)
    seen = set(headers)
    for it in items:
        if not isinstance(it, dict):
            return ParseResult(error="every entry must be an object")
        for k in it.keys():
            if k not in seen:
                seen.add(k)
                headers.append(str(k))

    rows: list[dict[str, str]] = []
    for it in items:
        row = {h: _coerce(it.get(h)) for h in headers}
        if _is_blank_row(row):
            continue
        rows.append(row)

    return ParseResult(headers=headers, rows=rows, row_count=len(rows))


def _parse_xlsx(content: bytes) -> ParseResult:
    # Lazy import — keeps `openpyxl` out of the import graph for tests
    # that don't need it.
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        return ParseResult(error="workbook has no sheets")

    headers: list[str] = []
    rows: list[dict[str, str]] = []
    for i, raw in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [_coerce(c) for c in raw]
            # Drop empty trailing header cells (Excel keeps them as None)
            while headers and headers[-1] == "":
                headers.pop()
            continue
        if len(rows) >= MAX_ROWS:
            return ParseResult(
                error=f"too many rows: > {MAX_ROWS}"
            )
        cells = [_coerce(c) for c in raw]
        # Pad short rows / clip long ones to header width
        cells = cells[: len(headers)] + [""] * (len(headers) - len(cells))
        row = dict(zip(headers, cells))
        if _is_blank_row(row):
            continue
        rows.append(row)

    if not headers:
        return ParseResult(error="no header row")
    return ParseResult(headers=headers, rows=rows, row_count=len(rows))


def _detect_csv_delimiter(text: str) -> str:
    """Pick `;` or `,` based on which appears more in the first non-empty line.
    Sniffer is too eager about quoting edge cases — this is good enough."""
    first = ""
    for line in text.splitlines():
        if line.strip():
            first = line
            break
    if not first:
        return ","
    return ";" if first.count(";") > first.count(",") else ","


def _parse_csv(content: bytes) -> ParseResult:
    try:
        text = content.decode("utf-8-sig")  # tolerate BOM
    except UnicodeDecodeError:
        try:
            text = content.decode("cp1251")  # Bitrix24 / older Excel exports
        except UnicodeDecodeError:
            return ParseResult(error="encoding not utf-8 or cp1251")

    delimiter = _detect_csv_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [_coerce(h) for h in (reader.fieldnames or [])]
    if not headers:
        return ParseResult(error="no header row")

    rows: list[dict[str, str]] = []
    for raw in reader:
        if len(rows) >= MAX_ROWS:
            return ParseResult(
                error=f"too many rows: > {MAX_ROWS}"
            )
        row = {h: _coerce(raw.get(h)) for h in headers}
        if _is_blank_row(row):
            continue
        rows.append(row)

    return ParseResult(headers=headers, rows=rows, row_count=len(rows))


def _unwrap_leads(data: Any) -> Any:
    """Accept either `[{...}, {...}]` or `{"leads": [...]}`."""
    if isinstance(data, dict) and isinstance(data.get("leads"), list):
        return data["leads"]
    return data


def _parse_json(content: bytes) -> ParseResult:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return ParseResult(error="encoding not utf-8")
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        return ParseResult(error=f"invalid JSON: {exc.msg}")
    items = _unwrap_leads(data)
    return _from_dict_rows([], items if isinstance(items, list) else [])


def _parse_yaml(content: bytes) -> ParseResult:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return ParseResult(error="encoding not utf-8")
    try:
        data = yaml.safe_load(text)
    except yaml.YAMLError as exc:
        return ParseResult(error=f"invalid YAML: {exc}")
    items = _unwrap_leads(data)
    return _from_dict_rows([], items if isinstance(items, list) else [])
