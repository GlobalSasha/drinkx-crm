"""Export-format encoders — pure bytes-in / bytes-out functions.

`leads_to_rows` flattens Lead ORM objects into plain dicts keyed by
EXPORT_FIELDS. The format-specific encoders take the resulting list of
dicts and return bytes ready for the download endpoint.

`export_md_zip` is the exception — it works directly off the Lead
ORM objects so the markdown body can pull from `ai_data.company_profile`
even when `include_ai_brief=False` for the tabular formats.

Stdlib for everything except `openpyxl` (XLSX) and `yaml` (PyYAML),
both already pinned by earlier sprints.
"""
from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from datetime import datetime
from typing import Any, Iterable

import yaml


EXPORT_FIELDS: list[str] = [
    "company_name",
    "segment",
    "city",
    "email",
    "phone",
    "website",
    "inn",
    "deal_amount",
    "priority",
    "deal_type",
    "source",
    "tags_json",
    "stage_name",
    "assigned_user_email",
    "fit_score",
    "created_at",
]

AI_BRIEF_COLUMN = "ai_company_profile"


# ---------------------------------------------------------------------------
# Lead → row flattening
# ---------------------------------------------------------------------------

def _join_tags(tags: Any) -> str:
    """Lead.tags_json is a Python list — flatten to comma-separated string."""
    if not tags:
        return ""
    if isinstance(tags, list):
        return ", ".join(str(t) for t in tags if t)
    return str(tags)


def _isoformat(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _ai_company_profile(lead: Any) -> str:
    ai_data = getattr(lead, "ai_data", None) or {}
    if not isinstance(ai_data, dict):
        return ""
    return str(ai_data.get("company_profile") or "")


def leads_to_rows(
    leads: Iterable[Any],
    *,
    stage_lookup: dict[Any, str] | None = None,
    user_email_lookup: dict[Any, str] | None = None,
    include_ai_brief: bool = False,
) -> list[dict[str, Any]]:
    """Flatten Lead ORM objects to plain dicts. `stage_lookup` and
    `user_email_lookup` carry the resolved relations the export task
    pre-fetches — keeps this function ORM-stub-friendly for tests."""
    stage_lookup = stage_lookup or {}
    user_email_lookup = user_email_lookup or {}

    rows: list[dict[str, Any]] = []
    for lead in leads:
        row: dict[str, Any] = {
            "company_name": getattr(lead, "company_name", "") or "",
            "segment": getattr(lead, "segment", "") or "",
            "city": getattr(lead, "city", "") or "",
            "email": getattr(lead, "email", "") or "",
            "phone": getattr(lead, "phone", "") or "",
            "website": getattr(lead, "website", "") or "",
            "inn": getattr(lead, "inn", "") or "",
            "deal_amount": getattr(lead, "deal_amount", "") or "",
            "priority": getattr(lead, "priority", "") or "",
            "deal_type": getattr(lead, "deal_type", "") or "",
            "source": getattr(lead, "source", "") or "",
            "tags_json": _join_tags(getattr(lead, "tags_json", None)),
            "stage_name": stage_lookup.get(getattr(lead, "stage_id", None), "") or "",
            "assigned_user_email": user_email_lookup.get(
                getattr(lead, "assigned_to", None), ""
            )
            or "",
            "fit_score": _isoformat(getattr(lead, "fit_score", None)),
            "created_at": _isoformat(getattr(lead, "created_at", None)),
        }
        if include_ai_brief:
            row[AI_BRIEF_COLUMN] = _ai_company_profile(lead)
        rows.append(row)
    return rows


def _columns_for(rows: list[dict[str, Any]]) -> list[str]:
    """Stable column order: canonical fields first, then ai_company_profile
    if any row carries it. Avoids set-based ordering surprises."""
    cols = list(EXPORT_FIELDS)
    if rows and any(AI_BRIEF_COLUMN in r for r in rows):
        cols.append(AI_BRIEF_COLUMN)
    return cols


# ---------------------------------------------------------------------------
# Tabular encoders
# ---------------------------------------------------------------------------

def export_csv(rows: list[dict[str, Any]]) -> bytes:
    """UTF-8 with BOM so Excel double-click opens with proper Cyrillic.
    Empty rows still emit a header — keeps downstream parsers happy."""
    cols = _columns_for(rows)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({c: r.get(c, "") for c in cols})
    text = buf.getvalue()
    return b"\xef\xbb\xbf" + text.encode("utf-8")


def export_json(rows: list[dict[str, Any]]) -> bytes:
    return json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8")


def export_yaml(rows: list[dict[str, Any]]) -> bytes:
    return yaml.safe_dump(rows, allow_unicode=True, sort_keys=False).encode("utf-8")


def export_xlsx(rows: list[dict[str, Any]]) -> bytes:
    # Lazy import — keeps openpyxl out of the import graph for tests
    # that exercise other formats without the dep.
    from openpyxl import Workbook
    from openpyxl.styles import Font

    cols = _columns_for(rows)
    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover — Workbook() always has an active sheet
        ws = wb.create_sheet()
    ws.title = "Leads"

    bold = Font(bold=True)
    ws.append(cols)
    for cell in ws[1]:
        cell.font = bold

    for r in rows:
        ws.append([r.get(c, "") for c in cols])

    # Pretty-but-bounded column widths (10..50 chars).
    for col_idx, col_name in enumerate(cols, start=1):
        longest = max(
            (len(str(r.get(col_name, ""))) for r in rows),
            default=len(col_name),
        )
        width = max(10, min(50, max(longest, len(col_name)) + 2))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
            width
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Markdown ZIP — one .md per lead, packaged together
# ---------------------------------------------------------------------------

_SLUG_RE = re.compile(r"[^\w\-]+", re.UNICODE)


def _slugify(value: str) -> str:
    r"""Best-effort filesystem-safe slug. Cyrillic survives via \w+ in
    UNICODE mode; ZIP filenames support UTF-8 since spec 6.3.0 (2006).
    Truncated to 64 chars to dodge OS path-length traps on extract."""
    cleaned = _SLUG_RE.sub("-", (value or "").strip()).strip("-")
    cleaned = re.sub(r"-+", "-", cleaned)
    return cleaned[:64] or "lead"


def _md_filename(lead: Any) -> str:
    inn = (getattr(lead, "inn", "") or "").strip()
    company = getattr(lead, "company_name", "") or "lead"
    return f"{inn}.md" if inn else f"{_slugify(company)}.md"


def _md_body(
    lead: Any,
    *,
    stage_name: str = "",
    assigned_email: str = "",
) -> str:
    company = getattr(lead, "company_name", "") or "—"
    fit = getattr(lead, "fit_score", None)
    fit_repr = "—" if fit is None else str(fit)
    parts = [
        f"# {company}",
        "",
        f"**Сегмент:** {getattr(lead, 'segment', '') or '—'}",
        f"**Город:** {getattr(lead, 'city', '') or '—'}",
        f"**Стадия:** {stage_name or '—'}",
        f"**Fit Score:** {fit_repr}",
        f"**Email:** {getattr(lead, 'email', '') or '—'}",
        f"**Телефон:** {getattr(lead, 'phone', '') or '—'}",
        f"**Сайт:** {getattr(lead, 'website', '') or '—'}",
        f"**Менеджер:** {assigned_email or '—'}",
        "",
        "## AI Brief",
        "",
        _ai_company_profile(lead) or "—",
        "",
    ]
    return "\n".join(parts)


def export_md_zip(
    leads: Iterable[Any],
    *,
    stage_lookup: dict[Any, str] | None = None,
    user_email_lookup: dict[Any, str] | None = None,
) -> bytes:
    """One .md per lead, packaged into a ZIP. Filename collisions are
    rare (INN-keyed when available, slug-fallback otherwise) but we
    suffix with `-2`, `-3`, … if they happen so we don't drop leads."""
    stage_lookup = stage_lookup or {}
    user_email_lookup = user_email_lookup or {}
    used_names: dict[str, int] = {}

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for lead in leads:
            name = _md_filename(lead)
            if name in used_names:
                used_names[name] += 1
                stem, ext = name.rsplit(".", 1)
                name = f"{stem}-{used_names[name]}.{ext}"
            else:
                used_names[name] = 1

            body = _md_body(
                lead,
                stage_name=stage_lookup.get(getattr(lead, "stage_id", None), ""),
                assigned_email=user_email_lookup.get(
                    getattr(lead, "assigned_to", None), ""
                ),
            )
            zf.writestr(name, body.encode("utf-8"))

    return out.getvalue()


# ---------------------------------------------------------------------------
# Format dispatch
# ---------------------------------------------------------------------------

def content_type_for(format_value: str) -> str:
    return {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv; charset=utf-8",
        "json": "application/json",
        "yaml": "application/yaml",
        "md_zip": "application/zip",
    }.get(format_value, "application/octet-stream")


def file_extension_for(format_value: str) -> str:
    return {
        "xlsx": "xlsx",
        "csv": "csv",
        "json": "json",
        "yaml": "yaml",
        "md_zip": "zip",
    }.get(format_value, "bin")


__all__ = [
    "EXPORT_FIELDS",
    "AI_BRIEF_COLUMN",
    "leads_to_rows",
    "export_csv",
    "export_json",
    "export_yaml",
    "export_xlsx",
    "export_md_zip",
    "content_type_for",
    "file_extension_for",
]
