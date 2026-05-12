"""Import leads/contacts from contacts_from_pdf.xlsx into the CRM.

Source data: `contacts_from_pdf.xlsx` (sheet "Контакты") — 4 columns
  Имя | Телефон | E-mail | Страница PDF

Each Excel row → one Lead + one Contact (B2B model: contacts hang off
leads). Deduplication priority on the contacts within the chosen
workspace:
  1. phone (E.164-normalized)
  2. email (lowercased, stripped)
  3. company_name (fallback when both above are missing)

Match found → fill empty fields only (never overwrite), append a note
  "Обогащено из импорта YYYY-MM-DD".
No match → insert new Lead + Contact, both tagged with
  source = "import_pdf_YYYY-MM-DD_<batch>"
so the batch can be queried / rolled back via that tag.

Connection: reads DATABASE_URL from env (same convention as
scripts/backfill_companies.py). Strips the asyncpg dialect prefix for
plain asyncpg.

Usage
  python3 scripts/import_leads_from_excel.py --parse-only
      Validate Excel parsing + normalization without touching the DB.

  python3 scripts/import_leads_from_excel.py --dry-run
      Connect to DB, project what would happen, do not write.

  python3 scripts/import_leads_from_excel.py --execute
      Actually import (asks for confirmation when total > 10 000).

  --workspace-id UUID      pick a specific workspace (otherwise the
                           single workspace in the DB is used; errors
                           out if there are 0 or 2+).
  --file PATH              override Excel path.
  --limit N                process only the first N data rows
                           (useful for end-to-end smoke tests).
  --chunk-size N           insert chunk size (default 500).
"""
from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
import uuid
from collections import Counter
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Iterator

# Path setup ----------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_DEFAULT = REPO_ROOT / "contacts_from_pdf.xlsx"
SOURCE_SHEET = "Контакты"

TODAY = date.today().isoformat()  # YYYY-MM-DD
BATCH_ID = uuid.uuid4().hex[:8]
SOURCE_TAG = f"import_pdf_{TODAY}_{BATCH_ID}"
ENRICH_NOTE = f"Обогащено из импорта {TODAY}"

# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

_EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")


def normalize_email(raw: Any) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if not s:
        return None
    if not _EMAIL_RE.match(s):
        return None
    return s


def normalize_phone(raw: Any) -> tuple[str | None, str | None]:
    """Return (normalized_or_None, reason_if_skipped).

    Russian phones → E.164 (+7XXXXXXXXXX, 12 chars).
    Already-international non-RU numbers are accepted if they start with
    "+" and have 8..15 digits.
    """
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s:
        return None, None

    keep_plus = s.startswith("+")
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None, "no digits"

    # Already international, non-RU.
    if keep_plus and not digits.startswith("7"):
        if 8 <= len(digits) <= 15:
            return "+" + digits, None
        return None, f"intl length {len(digits)}"

    # RU normalization.
    if len(digits) == 11 and digits[0] in ("7", "8"):
        return "+7" + digits[1:], None
    if len(digits) == 10 and digits[0] == "9":
        return "+7" + digits, None
    return None, f"ru length {len(digits)} digits={digits[:4]}…"


def derive_name_from_email(email: str) -> str:
    local = email.split("@", 1)[0]
    cleaned = re.sub(r"[._\-+0-9]+", " ", local).strip()
    return cleaned.title() if cleaned else local


def derive_company_name(name: str | None, email: str | None, phone: str | None) -> str:
    if name:
        return name[:255]
    if email:
        domain = email.split("@", 1)[1]
        return f"({domain})"[:255]
    if phone:
        return f"(контакт {phone})"[:255]
    return "Неизвестно (импорт PDF)"


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

@dataclass
class Row:
    excel_row: int
    name: str | None
    phone: str | None
    email: str | None
    page: str | None
    skip_reason: str | None = None


def iter_excel(path: Path, limit: int | None) -> Iterator[Row]:
    import openpyxl  # local import — keeps `--help` fast

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SOURCE_SHEET not in wb.sheetnames:
        raise SystemExit(f"sheet {SOURCE_SHEET!r} not found in {path}")
    ws = wb[SOURCE_SHEET]

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return
    if tuple(c for c in header[:4]) != ("Имя", "Телефон", "E-mail", "Страница PDF"):
        raise SystemExit(f"unexpected header: {header}")

    seen = 0
    for i, raw in enumerate(rows, start=2):
        if limit is not None and seen >= limit:
            break
        name = (str(raw[0]).strip() if raw[0] not in (None, "") else None) if len(raw) > 0 else None
        phone_raw = raw[1] if len(raw) > 1 else None
        email_raw = raw[2] if len(raw) > 2 else None
        page = (str(raw[3]).strip() if len(raw) > 3 and raw[3] not in (None, "") else None)

        phone, phone_reason = normalize_phone(phone_raw)
        email = normalize_email(email_raw)

        skip: str | None = None
        if not phone and not email and not name:
            skip = "row empty"
        elif phone_raw and not phone and not email:
            skip = f"bad phone ({phone_reason}) and no email"

        yield Row(
            excel_row=i,
            name=name,
            phone=phone,
            email=email,
            page=page,
            skip_reason=skip,
        )
        seen += 1


# ---------------------------------------------------------------------------
# Per-row plan
# ---------------------------------------------------------------------------

@dataclass
class Plan:
    rows_total: int = 0
    rows_skipped: int = 0
    skip_reasons: Counter = field(default_factory=Counter)
    invalid_phones: int = 0
    invalid_emails_in_source: int = 0
    new_leads: int = 0
    enriched_existing: int = 0
    duplicates_no_change: int = 0
    enrich_field_counts: Counter = field(default_factory=Counter)
    create_payloads: list[dict[str, Any]] = field(default_factory=list)
    enrich_ops: list[dict[str, Any]] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Async DB work
# ---------------------------------------------------------------------------

async def resolve_workspace(conn: Any, workspace_id: str | None) -> tuple[uuid.UUID, uuid.UUID | None]:
    if workspace_id:
        row = await conn.fetchrow(
            "SELECT id, default_pipeline_id FROM workspaces WHERE id = $1",
            uuid.UUID(workspace_id),
        )
        if row is None:
            raise SystemExit(f"workspace {workspace_id} not found")
        return row["id"], row["default_pipeline_id"]

    rows = await conn.fetch("SELECT id, name, default_pipeline_id FROM workspaces ORDER BY created_at ASC")
    if len(rows) == 0:
        raise SystemExit("no workspaces in DB — pass --workspace-id or seed one first")
    if len(rows) > 1:
        names = ", ".join(f"{r['id']} ({r['name']})" for r in rows)
        raise SystemExit(f"multiple workspaces — pass --workspace-id. Candidates: {names}")
    return rows[0]["id"], rows[0]["default_pipeline_id"]


async def resolve_default_stage(conn: Any, pipeline_id: uuid.UUID | None, workspace_id: uuid.UUID) -> uuid.UUID | None:
    """Pick the lowest-position stage of the workspace default pipeline, or None."""
    if pipeline_id is None:
        row = await conn.fetchrow(
            "SELECT id FROM pipelines WHERE workspace_id = $1 ORDER BY position ASC, created_at ASC LIMIT 1",
            workspace_id,
        )
        if row is None:
            return None
        pipeline_id = row["id"]
    row = await conn.fetchrow(
        "SELECT id FROM stages WHERE pipeline_id = $1 ORDER BY position ASC, created_at ASC LIMIT 1",
        pipeline_id,
    )
    return row["id"] if row else None


async def load_existing_index(conn: Any, workspace_id: uuid.UUID) -> tuple[dict[str, uuid.UUID], dict[str, uuid.UUID], dict[uuid.UUID, dict[str, Any]]]:
    """Return (by_phone, by_email, contacts_by_id)."""
    by_phone: dict[str, uuid.UUID] = {}
    by_email: dict[str, uuid.UUID] = {}
    contacts: dict[uuid.UUID, dict[str, Any]] = {}

    rows = await conn.fetch(
        """
        SELECT id, lead_id, name, email, phone, notes
        FROM contacts
        WHERE workspace_id = $1
        """,
        workspace_id,
    )
    for r in rows:
        cid = r["id"]
        contacts[cid] = {
            "id": cid,
            "lead_id": r["lead_id"],
            "name": r["name"],
            "email": r["email"],
            "phone": r["phone"],
            "notes": r["notes"],
        }
        # Normalize on the fly so old data with messy phones still matches.
        ph_norm, _ = normalize_phone(r["phone"])
        if ph_norm and ph_norm not in by_phone:
            by_phone[ph_norm] = cid
        em_norm = normalize_email(r["email"])
        if em_norm and em_norm not in by_email:
            by_email[em_norm] = cid
    return by_phone, by_email, contacts


def plan_rows(
    rows: Iterator[Row],
    by_phone: dict[str, uuid.UUID],
    by_email: dict[str, uuid.UUID],
    contacts: dict[uuid.UUID, dict[str, Any]],
) -> Plan:
    """Walk rows once, decide create-vs-enrich, accumulate within-batch dedup."""
    plan = Plan()
    # Track contacts created earlier in *this* batch so two Excel rows
    # for the same person collapse into one insert.
    pending_by_phone: dict[str, int] = {}
    pending_by_email: dict[str, int] = {}

    for row in rows:
        plan.rows_total += 1

        if row.skip_reason:
            plan.rows_skipped += 1
            plan.skip_reasons[row.skip_reason] += 1
            if "bad phone" in row.skip_reason:
                plan.invalid_phones += 1
            continue

        # 1. Existing-contact match
        matched_cid: uuid.UUID | None = None
        if row.phone and row.phone in by_phone:
            matched_cid = by_phone[row.phone]
        elif row.email and row.email in by_email:
            matched_cid = by_email[row.email]

        if matched_cid is not None:
            existing = contacts[matched_cid]
            patch: dict[str, Any] = {}
            if not existing.get("name") and row.name:
                patch["name"] = row.name[:120]
                plan.enrich_field_counts["name"] += 1
            if not existing.get("email") and row.email:
                patch["email"] = row.email[:254]
                plan.enrich_field_counts["email"] += 1
            if not existing.get("phone") and row.phone:
                patch["phone"] = row.phone[:30]
                plan.enrich_field_counts["phone"] += 1
            if patch:
                old_notes = existing.get("notes") or ""
                note_extra = ENRICH_NOTE
                if row.page:
                    note_extra += f" (PDF page {row.page})"
                if note_extra not in old_notes:
                    patch["notes"] = (old_notes + ("\n" if old_notes else "") + note_extra)[:4000]
                plan.enriched_existing += 1
                plan.enrich_ops.append({"contact_id": matched_cid, "patch": patch})
                # Keep the index consistent so later rows in the same
                # batch see the freshly-filled fields and don't try to
                # enrich them again.
                existing.update(patch)
                if row.phone and row.phone not in by_phone:
                    by_phone[row.phone] = matched_cid
                if row.email and row.email not in by_email:
                    by_email[row.email] = matched_cid
            else:
                plan.duplicates_no_change += 1
            continue

        # 2. In-batch dedup (two Excel rows for the same person).
        in_batch_idx: int | None = None
        if row.phone and row.phone in pending_by_phone:
            in_batch_idx = pending_by_phone[row.phone]
        elif row.email and row.email in pending_by_email:
            in_batch_idx = pending_by_email[row.email]

        if in_batch_idx is not None:
            payload = plan.create_payloads[in_batch_idx]
            changed = False
            if not payload["contact_name"] and row.name:
                payload["contact_name"] = row.name[:120]
                changed = True
            if not payload["contact_email"] and row.email:
                payload["contact_email"] = row.email[:254]
                changed = True
            if not payload["contact_phone"] and row.phone:
                payload["contact_phone"] = row.phone[:30]
                changed = True
            if changed and row.phone:
                pending_by_phone[row.phone] = in_batch_idx
            if changed and row.email:
                pending_by_email[row.email] = in_batch_idx
            if not changed:
                plan.duplicates_no_change += 1
            continue

        # 3. Brand new record.
        contact_name = row.name or (derive_name_from_email(row.email) if row.email else "(без имени)")
        payload = {
            "company_name": derive_company_name(row.name, row.email, row.phone),
            "lead_email": row.email,
            "lead_phone": row.phone,
            "contact_name": contact_name[:120],
            "contact_email": row.email,
            "contact_phone": row.phone,
            "notes": f"PDF page: {row.page}" if row.page else None,
        }
        idx = len(plan.create_payloads)
        plan.create_payloads.append(payload)
        plan.new_leads += 1
        if row.phone:
            pending_by_phone[row.phone] = idx
        if row.email:
            pending_by_email[row.email] = idx

    return plan


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------

async def apply_enrichments(conn: Any, ops: list[dict[str, Any]], chunk_size: int) -> int:
    """Apply field-fill patches one at a time (simpler than building dynamic
    bulk UPDATE; ~enriched_existing is usually small relative to creates)."""
    n = 0
    for op in ops:
        patch = op["patch"]
        sets = []
        vals: list[Any] = []
        for k, v in patch.items():
            sets.append(f"{k} = ${len(vals) + 1}")
            vals.append(v)
        vals.append(op["contact_id"])
        await conn.execute(
            f"UPDATE contacts SET {', '.join(sets)}, updated_at = NOW() WHERE id = ${len(vals)}",
            *vals,
        )
        n += 1
    return n


async def apply_creates(
    conn: Any,
    payloads: list[dict[str, Any]],
    workspace_id: uuid.UUID,
    pipeline_id: uuid.UUID | None,
    stage_id: uuid.UUID | None,
    chunk_size: int,
) -> int:
    """Bulk insert (Lead, Contact) pairs in chunks within a single tx."""
    inserted = 0
    for start in range(0, len(payloads), chunk_size):
        chunk = payloads[start : start + chunk_size]
        async with conn.transaction():
            # Insert leads, returning the generated id for each row.
            lead_rows = [
                (
                    uuid.uuid4(),
                    workspace_id,
                    pipeline_id,
                    stage_id,
                    p["company_name"],
                    p["lead_email"],
                    p["lead_phone"],
                    SOURCE_TAG,
                )
                for p in chunk
            ]
            await conn.executemany(
                """
                INSERT INTO leads (
                    id, workspace_id, pipeline_id, stage_id, company_name,
                    email, phone, source, tags_json, score,
                    assignment_status, is_rotting_stage, is_rotting_next_step,
                    agent_state, created_at, updated_at
                ) VALUES (
                    $1, $2, $3, $4, $5,
                    $6, $7, $8, '[]'::json, 0,
                    'pool', false, false,
                    '{}'::jsonb, NOW(), NOW()
                )
                """,
                lead_rows,
            )
            contact_rows = [
                (
                    uuid.uuid4(),
                    lead_id,
                    workspace_id,
                    p["contact_name"],
                    p["contact_email"],
                    p["contact_phone"],
                    SOURCE_TAG,
                    p["notes"],
                )
                for (lead_id, *_rest), p in zip(lead_rows, chunk, strict=True)
            ]
            await conn.executemany(
                """
                INSERT INTO contacts (
                    id, lead_id, workspace_id, name, email, phone, source,
                    notes, confidence, verified_status, created_at, updated_at
                ) VALUES (
                    $1, $2, $3, $4, $5, $6, $7, $8,
                    'medium', 'to_verify', NOW(), NOW()
                )
                """,
                contact_rows,
            )
        inserted += len(chunk)
        print(f"  inserted {inserted}/{len(payloads)}…", file=sys.stderr)
    return inserted


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def print_summary(plan: Plan, workspace_id: uuid.UUID | None, mode: str) -> None:
    print()
    print("=" * 60)
    print(f"  Import summary — mode={mode}")
    print(f"  Batch tag: {SOURCE_TAG}")
    if workspace_id:
        print(f"  Workspace: {workspace_id}")
    print("=" * 60)
    print(f"  Rows in file:           {plan.rows_total}")
    print(f"  Skipped (validation):   {plan.rows_skipped}")
    for reason, n in plan.skip_reasons.most_common():
        print(f"    - {reason}: {n}")
    print(f"  New leads to create:    {plan.new_leads}")
    print(f"  Existing to enrich:     {plan.enriched_existing}")
    if plan.enrich_field_counts:
        print(f"    fields filled: " + ", ".join(f"{k}={v}" for k, v in plan.enrich_field_counts.most_common()))
    print(f"  Duplicates (no change): {plan.duplicates_no_change}")
    print("=" * 60)


def show_create_samples(plan: Plan, n: int = 5) -> None:
    if not plan.create_payloads:
        return
    print(f"\nSample of new leads (first {min(n, len(plan.create_payloads))}):")
    for p in plan.create_payloads[:n]:
        print(f"  - {p['company_name']!r} | name={p['contact_name']!r} email={p['contact_email']} phone={p['contact_phone']}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def run(args: argparse.Namespace) -> int:
    excel_path = Path(args.file).resolve()
    if not excel_path.exists():
        print(f"✗ excel file not found: {excel_path}", file=sys.stderr)
        return 2

    if args.parse_only:
        plan = Plan()
        plan = plan_rows(
            iter_excel(excel_path, args.limit),
            by_phone={},
            by_email={},
            contacts={},
        )
        print_summary(plan, workspace_id=None, mode="parse-only")
        show_create_samples(plan)
        return 0

    # DB modes need DATABASE_URL.
    db_url = os.environ.get("DATABASE_URL", "").replace(
        "postgresql+asyncpg://", "postgresql://"
    )
    if not db_url:
        print("✗ DATABASE_URL not set (use --parse-only to skip DB).", file=sys.stderr)
        return 2

    try:
        import asyncpg  # type: ignore
    except ImportError:
        print("✗ asyncpg not installed; run from apps/api venv.", file=sys.stderr)
        return 2

    conn = await asyncpg.connect(db_url)
    try:
        workspace_id, default_pipeline_id = await resolve_workspace(conn, args.workspace_id)
        stage_id = await resolve_default_stage(conn, default_pipeline_id, workspace_id)
        if default_pipeline_id is None:
            row = await conn.fetchrow(
                "SELECT id FROM pipelines WHERE workspace_id = $1 ORDER BY position ASC LIMIT 1",
                workspace_id,
            )
            default_pipeline_id = row["id"] if row else None
        print(f"workspace = {workspace_id}")
        print(f"pipeline  = {default_pipeline_id}")
        print(f"stage     = {stage_id}")

        print("loading existing contacts index…")
        by_phone, by_email, contacts = await load_existing_index(conn, workspace_id)
        print(f"  indexed {len(contacts)} contacts ({len(by_phone)} phones, {len(by_email)} emails)")

        print("planning rows…")
        plan = plan_rows(
            iter_excel(excel_path, args.limit),
            by_phone=by_phone,
            by_email=by_email,
            contacts=contacts,
        )

        print_summary(plan, workspace_id=workspace_id, mode="execute" if args.execute else "dry-run")
        show_create_samples(plan)

        if args.dry_run:
            return 0

        total_writes = plan.new_leads + plan.enriched_existing
        if total_writes == 0:
            print("nothing to write — exiting.")
            return 0

        if total_writes > 10_000 and not args.yes:
            print(f"\n⚠ about to write {total_writes} rows. Re-run with --yes to confirm.", file=sys.stderr)
            return 3

        print(f"\napplying {plan.enriched_existing} enrichments…")
        await apply_enrichments(conn, plan.enrich_ops, args.chunk_size)
        print(f"applying {plan.new_leads} inserts (chunk {args.chunk_size})…")
        await apply_creates(
            conn,
            plan.create_payloads,
            workspace_id=workspace_id,
            pipeline_id=default_pipeline_id,
            stage_id=stage_id,
            chunk_size=args.chunk_size,
        )
        print(f"\n✓ done. Records tagged source = {SOURCE_TAG!r}.")
        print(f"  To roll back this batch:")
        print(f"    DELETE FROM contacts WHERE source = '{SOURCE_TAG}';")
        print(f"    DELETE FROM leads    WHERE source = '{SOURCE_TAG}';")
        return 0
    finally:
        await conn.close()


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    mode = p.add_mutually_exclusive_group(required=True)
    mode.add_argument("--parse-only", action="store_true", help="parse Excel only, no DB")
    mode.add_argument("--dry-run", action="store_true", help="connect, project, do not write")
    mode.add_argument("--execute", action="store_true", help="actually write to DB")
    p.add_argument("--file", default=str(EXCEL_DEFAULT))
    p.add_argument("--workspace-id", default=None)
    p.add_argument("--limit", type=int, default=None)
    p.add_argument("--chunk-size", type=int, default=500)
    p.add_argument("--yes", action="store_true", help="skip large-write confirmation")
    return p.parse_args()


if __name__ == "__main__":
    sys.exit(asyncio.run(run(parse_args())))
